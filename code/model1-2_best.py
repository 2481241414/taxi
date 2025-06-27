#coding:utf-8
from openpyxl import load_workbook
import requests
import json
import time
import random
import string
import hmac
import hashlib
import base64
from datetime import datetime, timezone

excel_name = r"D:/工作内容/202505/打车语料-优化版.xlsx"

def read_excel(i,k):
    wb = load_workbook(excel_name)
    sheetname = wb.sheetnames
    sheet = wb[sheetname[k]]
    query = sheet.cell(row=i,column=3).value
    if sheet.cell(row=i,column=4).value == "滴滴车列表":
         extInput = sheet.cell(row=2,column=2).value
    elif sheet.cell(row=i,column=4).value == "高德车列表":
         extInput = sheet.cell(row=3,column=2).value
    else:
         extInput = sheet.cell(row=3,column=2).value
    sysSessionId = sheet.cell(row=i,column=5).value
    return query,extInput,sysSessionId

def read_excel_duolun(i,k):
    wb = load_workbook(excel_name)
    sheetname = wb.sheetnames
    sheet = wb[sheetname[k]]
    query1 = sheet.cell(row=i,column=3).value
    query2 = sheet.cell(row=i,column=4).value
    query3 = sheet.cell(row=i,column=5).value
    if sheet.cell(row=i,column=6).value == "滴滴车列表":
         extInput = sheet.cell(row=2,column=2).value
    elif sheet.cell(row=i,column=6).value == "高德车列表":
         extInput = sheet.cell(row=3,column=2).value
    else:
         extInput = sheet.cell(row=3,column=2).value
    sysSessionId = sheet.cell(row=i,column=7).value
    return query1,query2,query3,extInput,sysSessionId

def write_excel(i,k,model1_result, content, content2, model2_result):
    wb = load_workbook(excel_name)
    sheetname = wb.sheetnames
    sheet = wb[sheetname[k]]
    sheet.cell(i,6).value = str(model1_result)
    sheet.cell(i,7).value = str(content)
    sheet.cell(i,8).value = str(content2)
    sheet.cell(i,10).value = str(model2_result)
    wb.save(excel_name)

def write_excel_duolun(i,k,model1_result, content, content2, model2_result):
    wb = load_workbook(excel_name)
    sheetname = wb.sheetnames
    sheet = wb[sheetname[k]]
    sheet.cell(i,8).value = str(model1_result)
    sheet.cell(i,9).value = str(content)
    sheet.cell(i,10).value = str(content2)
    sheet.cell(i,12).value = str(model2_result)
    wb.save(excel_name)

def generate_hmac_headers():
    # 密钥配置
    ACCESS_KEY = "aip_yoyo_assistant_1744015109999"
    SECRET_KEY = "iTN%T#%g42IQk#Y-g987z-QCE@L@vUHf45a".encode('utf-8')

    # 1. 生成时间戳
    date_header = datetime.now(timezone.utc).strftime('%a, %d %b %Y %H:%M:%S GMT')

    # 2. 构建签名头列表
    signed_headers = [
        ("X-HMAC-ALGORITHM", "hmac-sha256"),
        ("X-HMAC-ACCESS-KEY", ACCESS_KEY),
        ("x-request-nonce", "20250428lhx003")
    ]

    # 3. 构造签名字符串
    method = "POST"
    path = "/aip/access-agent/cloud-api/v1/task/take_taxi_strategy_analyze/aip_yoyo_assistant"
    params = ""  # 根据实际请求参数设置

    signing_components = [
        method,
        path,
        params,
        ACCESS_KEY,
        date_header,
        "\n".join([f"{k}:{v}" for k, v in signed_headers])
    ]
    signing_string = "\n".join(signing_components) + "\n"  # 末尾必须换行

    # 4. 计算签名
    signature = hmac.new(
        SECRET_KEY,
        signing_string.encode('utf-8'),
        hashlib.sha256
    ).digest()
    base64_signature = base64.b64encode(signature).decode('utf-8')

    # 5. 组装请求头
    headers = {
        "X-HMAC-ACCESS-KEY": ACCESS_KEY,
        "X-HMAC-SIGNATURE": base64_signature,
        "X-HMAC-ALGORITHM": "hmac-sha256",
        "Date": date_header,
        "X-HMAC-SIGNED-HEADERS": ";".join([k for k, v in signed_headers]),
        "x-request-nonce": "20250428lhx003",
        'x-device-type': 'PHONE',
        'x-country': 'CN',
        'x-rom-version': '234',
        'x-language': 'zh',
        'x-time-zone': '234',
        'x-model': 'ATQ',
        'Content-Type': 'application/json',
        'x-req-env': 'gray'
    }

    return headers

def generate_hmac_headers1():
    # 密钥配置
    ACCESS_KEY = "aip_yoyo_assistant_1744015109999"
    SECRET_KEY = "iTN%T#%g42IQk#Y-g987z-QCE@L@vUHf45a".encode('utf-8')

    # 1. 生成时间戳
    date_header = datetime.now(timezone.utc).strftime('%a, %d %b %Y %H:%M:%S GMT')

    # 2. 构建签名头列表
    signed_headers = [
        ("X-HMAC-ALGORITHM", "hmac-sha256"),
        ("X-HMAC-ACCESS-KEY", ACCESS_KEY),
        ("x-request-nonce", "20250429lhx002")
    ]

    # 3. 构造签名字符串
    method = "POST"
    path = "/aip/access-agent/cloud-api/v1/task/take_taxi_sku_choose/aip_yoyo_assistant"
    params = ""  # 根据实际请求参数设置

    signing_components = [
        method,
        path,
        params,
        ACCESS_KEY,
        date_header,
        "\n".join([f"{k}:{v}" for k, v in signed_headers])
    ]
    signing_string = "\n".join(signing_components) + "\n"  # 末尾必须换行

    # 4. 计算签名
    signature = hmac.new(
        SECRET_KEY,
        signing_string.encode('utf-8'),
        hashlib.sha256
    ).digest()
    base64_signature = base64.b64encode(signature).decode('utf-8')

    # 5. 组装请求头
    headers = {
        "X-HMAC-ACCESS-KEY": ACCESS_KEY,
        "X-HMAC-SIGNATURE": base64_signature,
        "X-HMAC-ALGORITHM": "hmac-sha256",
        "Date": date_header,
        "X-HMAC-SIGNED-HEADERS": ";".join([k for k, v in signed_headers]),
        'x-request-nonce': '20250429lhx002',
        'x-device-type': 'PHONE',
        'x-country': 'CN',
        'x-rom-version': '234',
        'x-language': 'zh',
        'x-time-zone': '234',
        'x-model': 'ATQ',
        'Content-Type': 'application/json',
        'x-req-env': 'gray',
        # 'x-origin-udid': 'wuchengtest'
        # 'x-origin-udid': 'betacase0604faqfzy'
    }
    return headers

# 模型1
def getresult(query,sysSessionId):
    url = "http://aiplatac-pre-drcn.inner.cloud.hihonor.com/aip/access-agent/cloud-api/v1/task/take_taxi_strategy_analyze/aip_yoyo_assistant"

    payload = json.dumps({
        "session": {
            "sessionId": "{}".format(sysSessionId),
            "sysSessionId": "{}".format(sysSessionId)
        },
        "input": {
            "messageList": [
                {
                    "role": "user",
                    "content": "{}".format(query)
                }
            ]
        },
        "switches": [
            {
                "key": "developerDebugMode",
                "value": "on"
            }
        ]
    })

    headers = generate_hmac_headers()

    response = requests.request("POST", url, headers=headers,data=payload)

    return json.loads(response.content)["data"]["taskResult"]["strategyList"][0]

def getresult_duolun(query1, query2, query3, sysSessionId):
    url = "http://aiplatac-pre-drcn.inner.cloud.hihonor.com/aip/access-agent/cloud-api/v1/task/take_taxi_strategy_analyze/aip_yoyo_assistant"

    payload = json.dumps({
        "session": {
            "sessionId": "{}".format(sysSessionId),
            "sysSessionId": "{}".format(sysSessionId)
        },
        "input": {
            "messageList": [
                {
                    "content": "{}".format(query1),
                    "role": "user"
                },
                {
                    "content": "好的，请进一步描述你的需求",
                    "role": "assistant"
                },
                {
                    "content": "{}".format(query2),
                    "role": "user"
                },
                {
                    "content": "好的，请进一步描述你的需求",
                    "role": "assistant"
                },
                {
                    "content": "{}".format(query3),
                    "role": "user"
                }
            ]
        },
        "switches": [
            {
                "key": "developerDebugMode",
                "value": "on"
            }
        ]
    })

    headers = generate_hmac_headers()

    response = requests.request("POST", url, headers=headers,data=payload)

    return json.loads(response.content)["data"]["taskResult"]["strategyList"][0]

def get_model1_content(sysSessionId):
    url = "https://wo-drcn-test.cloud.hihonor.com/aip-pre/aip-admin/v1/message/search"

    headers = {
        'Content-Type': 'application/json',
        'cookie': 'Path=/; portal=inter; lang=zh; woLang=zh; honor_wp_lang=zh_CN; deviceId=1daf8ec30b75f04062081810443bed13; TID=1daf8ec30b75f04062081810443bed13; GLOBAL_SESSION_CMP_BETA=fe081db8-4c06-4dc6-8a3c-0f2573ff5fcc; logFlag=in; suid=87C27D6ADC5D1B23DBDAD4BF826DC783EB04498E44B4D62893A09DF729555E59; hwssot3=31357001167422; uid=313031303AC424DBD9D9C200E98158A2B56596A383861197DBAEE5EAB530520967C2FD59A84EB127DA2F5328EE; hwssotinter3=31357575015620; authmethod=31303130F0D47F8279C633D4FB52C79FB6829D58F1793E0521AA89357B953766319178C7694C; hwssotinter=31303130FFEAA7E091453EEDF61569B76A96F7E9DB2250FFD7C1D5E2C6849CC515447AAAB55602D0D1FD02C5CBA0D9DCAC; hwsso_uniportal=B00aA56MvAidEuYBEqUNbm6ZEEFIFadGNhNWB93qZHbCjJB9Nh0ncAyC1_bPekILDpuc4cupbdh6rKWNmt0R0GqS5I5z_b_bgDX0Uhmglul_bjHL5181v02T0nggoFozPby7tFHmQrMl6M1if2Oeh1N029ivGMqx6sWULu5Rdk4bQOjjMSlYojfhH39_aQeyMiokBYL4_bLSfdkq5GZN32cIlZd621rwfh4O_bfnOIXdpUwFIvPqKvUEZCR04pLpUhajFJ7cAgDX5ZA_b7GdkpL3V3FIqZ2Q1ZsSvwzzgY2Rnu6NR1TOSuTUtYks_aC2g0t_b_a1RD6xaO5oWIVjsMSXdlyJ6cN6w_c_c####jDrTpRCfxdjaiP7GVQ8C4iCuqPDPmVgL40_ajsvBqg63UpbY3urxR_bzNUw8zbsExD2BCQ4AqXvtaZwOWS_a10g5MfT2qFMUrWA9eaaO5RsfLSYRXyeGJ4F9yMeXPXPRqSF8eUnelrn4gElNwdHCiq9KOfYniya1EdTg2B1qwm51_bw1gk8swBvAtKNlzeU3gfumPWkHiHg1RFnsuqwbIED11jLlu7nYtsqaIhBawaaAGN8JUZ1077cLpT2XrJr2_bB5XE_buzUkzS1er4lofxEuU6WHNyHLmCeujAVOghbw4ONP_asyUBxgcXGBWDz_b4QjPGLgMcFpjiiIePHdJIvR0TFTeg_c_c####QrXbRoAERF2ABVdrtHCQ0zIuT9X61ABsbOWOlE9wPETAJp7OU1VuyxA8DhCHNosbBAqqtuwKsQ1G9PELWbn_aTGJjn7y4jFuBsOvGAKCG9Yxx3nyXMn9bjzAdvnQ_aVJGG9sjktAVa9P011kPDOJKqwdlCNl0GAqW5Kcx9uiK6lYErr7hEfg7WQruXYW81SMTQSCBFYp0XSrIvjqkhTYL35sCFjrQyyeRoBTztBvNleG5wFhUjjNEjQqFiQmBo4gjuuFpj_a6LADSW9tI5_aUF42QfLsq3CwlK02zTY71uoFm2JOjTHTASVl0YzWVNMJdrOTrwoHOUw0hK1p_aUWL8KC0QA_c_c; sip=3130313068554FEB83CE1CD915E487FB5E2061BCB0AB58397FD64739BA0091B7C6B695CD0552DC01DE194962BF5245B53B608A2A6A54D8629538E15927BF1DBDCDE2A53DEC14C04D156B7A13871EE097917772A8416762831615A144; sid=3130313035124D6A070F745BACC715BF1AE6A0753D20DCA1D45BC7B76E78BAD233C037DE30186C4504CBFB63A5E542F988EBF4B61172572351449D455D6DB4B7EFC519E7D0EB12B4; woAuth=QUVTR0NNQEREQzlDQURFM0E5Rjk2NUMyNUMwMzBERjoyNEMzQjMwOUZBMzkwMzJFMjk1ODFGQzRDRDFGRjM4OEU5QTBGNTg0Mjk3NUFGREJGQTc3Mjk2ODdEQjczMkM5MDhGOTFERUIxRkI2NTcxRTkyMDc2NDgyRUU3NkI3QjYxMjgzOUI2QzNCMDk2QkE2OEJBQzZFNEREOThFRTU5MA=='
    }

    data = {"sysSessionId": sysSessionId, "clientId": "", "producerNodeType": "", "currentRoundId": "", "recentRounds": 5}

    # 创建一个Session对象
    session = requests.Session()

    # 发送POST请求
    response = session.post(url, json=data, headers=headers, stream=True)

    if response.status_code == 200:
        result = response.content.decode('utf-8')
        try:
            json_data = json.loads(result)
            if len(json_data.get("data", [])) >= 2:
                content = json_data["data"][-2]['messageDetail']['content']
                content2 = json_data["data"][-1]['messageDetail']['content']
            else:
                content = ''
                content2 = ''
        except json.JSONDecodeError as e:
            print(f"JSON解析错误: {e}")
    else:
        # 如果请求失败，写入错误信息
        content = f"Error: {response.status_code}"
        content2 = f"Error: {response.status_code}"

    return content,content2

# 模型2
def getresult2(query, extInput, sysSessionId):
    url = "http://aiplatac-pre-drcn.inner.cloud.hihonor.com/aip/access-agent/cloud-api/v1/task/take_taxi_sku_choose/aip_yoyo_assistant"

    payload = json.dumps({
        "session": {
            "sessionId": "{}".format(sysSessionId),
            "sysSessionId": "{}".format(sysSessionId)
        },
        "input": {
            "messageList": [
                {
                    "role": "user",
                    "content": "{}".format(query),
                    "taskExts": [
                        {
                            "extType": "take_taxi_sku_choose",
                            "extInput": json.loads(extInput)
                        }
                    ]
                }
            ]
        }
    })
    headers = generate_hmac_headers1()

    response = requests.request("POST", url, headers=headers,data=payload)

    return json.loads(response.content)["data"]["taskResult"]["selectList"]

def getresult2_duolun(query1, query2, query3, extInput, sysSessionId):
    url = "http://aiplatac-pre-drcn.inner.cloud.hihonor.com/aip/access-agent/cloud-api/v1/task/take_taxi_sku_choose/aip_yoyo_assistant"

    payload = json.dumps({
        "session": {
            "sessionId": "{}".format(sysSessionId),
            "sysSessionId": "{}".format(sysSessionId)
        },
        "input": {
            "messageList": [{
                    "content": "{}".format(query1),
                    "role": "user"
                },
                {
                    "content": "好的，请进一步描述你的需求",
                    "role": "assistant"
                },
                {
                    "content": "{}".format(query2),
                    "role": "user"
                },
                {
                    "content": "好的，请进一步描述你的需求",
                    "role": "assistant"
                },
                {
                    "content": "{}".format(query3),
                    "role": "user",
                    "taskExts": [
                        {
                            "extType": "take_taxi_sku_choose",
                            "extInput":  json.loads(extInput)
                        }
                    ]
                }
            ]
        }
    })

    headers = generate_hmac_headers1()

    response = requests.request("POST", url, headers=headers,data=payload)

    return json.loads(response.content)["data"]["taskResult"]["selectList"]

if __name__ == '__main__':

    # 价格推理
    for i in range(174,202):
        print("价格推理" + str(i))
        query, extInput, sysSessionId = read_excel(i,0)
        model1_result = getresult(query,sysSessionId)
        print(model1_result)
        time.sleep(2)
        content, content2 = get_model1_content(sysSessionId)
        print(content)
        print(content2)
        model2_result = getresult2(query, extInput, sysSessionId)
        print(model2_result)
        write_excel(i, 0, model1_result, content, content2, model2_result)
    #
    # # 车型理解
    # for i in range(2,202):
    #     print("车型理解" + str(i))
    #     query, extInput, sysSessionId = read_excel(i,1)
    #     model1_result = getresult(query,sysSessionId)
    #     print(model1_result)
    #     time.sleep(2)
    #     content, content2 = get_model1_content(sysSessionId)
    #     print(content)
    #     print(content2)
    #     model2_result = getresult2(query, extInput, sysSessionId)
    #     print(model2_result)
    #     write_excel(i, 1, model1_result, content, content2, model2_result)
    #
    # # 组合场景
    # for i in range(2,377):
    #     print("组合场景" + str(i))
    #     query, extInput, sysSessionId = read_excel(i,2)
    #     model1_result = getresult(query,sysSessionId)
    #     print(model1_result)
    #     time.sleep(2)
    #     content, content2 = get_model1_content(sysSessionId)
    #     print(content)
    #     print(content2)
    #     model2_result = getresult2(query, extInput, sysSessionId)
    #     print(model2_result)
    #     write_excel(i, 2, model1_result, content, content2, model2_result)
    #
    # # appName场景
    # for i in range(2,92):
    #     print("appName场景" + str(i))
    #     query, extInput, sysSessionId = read_excel(i,3)
    #     model1_result = getresult(query,sysSessionId)
    #     print(model1_result)
    #     time.sleep(2)
    #     content, content2 = get_model1_content(sysSessionId)
    #     print(content)
    #     print(content2)
    #     model2_result = getresult2(query, extInput, sysSessionId)
    #     print(model2_result)
    #     write_excel(i, 3, model1_result, content, content2, model2_result)

    # 多轮对话
    for i in range(49,113):
        print("多轮对话" + str(i))
        query1, query2, query3, extInput, sysSessionId = read_excel_duolun(i,4)
        model1_result = getresult_duolun(query1, query2, query3, sysSessionId)
        print(model1_result)
        time.sleep(2)
        content, content2 = get_model1_content(sysSessionId)
        print(content)
        print(content2)
        model2_result = getresult2_duolun(query1, query2, query3, extInput, sysSessionId)
        print(model2_result)
        write_excel_duolun(i, 4, model1_result, content, content2, model2_result)